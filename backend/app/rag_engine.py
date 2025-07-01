"""
🧠 Enhanced KI-QMS RAG Engine - Retrieval Augmented Generation with OCR

Chat mit allen QMS-Dokumenten inklusive Bildern und Diagrammen:
- "Welche Schraube bei Antriebseinheit verwenden?"
- "Was zeigt das Flussdiagramm in der SOP?"
- "Wie funktioniert der Kalibrierungsprozess?"

Enhanced Features:
- 🔍 OCR für Bilder und Diagramme
- 🏠 Lokale Vector Database (ChromaDB) 
- 🆓 Kostenlose Embeddings (sentence-transformers)
- 🌟 Google Gemini für Chat
- 📚 Erweiterte Dokumentverarbeitung (PDF, DOCX, Excel, Bilder)
- 🎯 Quellenangaben mit Content-Type

Autoren: KI-QMS Team
Version: 2.0.0
"""

import os
import logging
from typing import List, Optional, Dict, Any
from dataclasses import dataclass
from datetime import datetime
import sqlite3

# Core dependencies
try:
    import chromadb
    from sentence_transformers import SentenceTransformer
    import google.generativeai as genai
    CORE_AVAILABLE = True
except ImportError as e:
    logging.error(f"Core RAG dependencies missing: {e}")
    CORE_AVAILABLE = False

# OCR and document processing dependencies
try:
    import pytesseract
    from PIL import Image
    import cv2
    import numpy as np
    from PyPDF2 import PdfReader
    from docx import Document
    import openpyxl
    from pdf2image import convert_from_path
    OCR_AVAILABLE = True
except ImportError as e:
    logging.warning(f"OCR dependencies not available: {e}")
    OCR_AVAILABLE = False

@dataclass
class DocumentChunk:
    """Represents a chunk of processed document content."""
    content: str
    metadata: Dict[str, Any]
    chunk_id: str
    source_file: str
    chunk_index: int
    content_type: str = "text"  # text, image, table, diagram

@dataclass 
class RAGResult:
    """Result from RAG query with enhanced metadata."""
    response: str
    sources: List[Dict[str, Any]]
    confidence: float
    query_embedding: Optional[List[float]] = None

class EnhancedDocumentProcessor:
    """Advanced document processor with OCR capabilities."""
    
    def __init__(self):
        self.ocr_available = OCR_AVAILABLE
        if self.ocr_available:
            self.setup_tesseract()
    
    def setup_tesseract(self):
        """Setup Tesseract OCR configuration."""
        if not self.ocr_available:
            return
        self.tesseract_config = r'--oem 3 --psm 6 -l deu+eng'
        
    def extract_text_from_image(self, image_path: str) -> str:
        """Extract text from image using OCR."""
        if not self.ocr_available:
            return ""
            
        try:
            # Load and preprocess image
            image = cv2.imread(image_path)
            if image is None:
                return ""
                
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Apply image enhancement for better OCR
            gray = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
            gray = cv2.medianBlur(gray, 3)
            
            # Extract text
            text = pytesseract.image_to_string(gray, config=self.tesseract_config)
            return text.strip()
        except Exception as e:
            logging.warning(f"OCR failed for {image_path}: {e}")
            return ""
    
    def extract_images_from_pdf(self, pdf_path: str) -> List[Dict[str, Any]]:
        """Extract images from PDF for OCR processing."""
        if not self.ocr_available:
            return []
            
        images_data = []
        try:
            # Convert PDF pages to images
            images = convert_from_path(pdf_path, dpi=300)
            
            for i, image in enumerate(images):
                # Save temporary image
                temp_path = f"/tmp/pdf_page_{i}.png"
                image.save(temp_path, 'PNG')
                
                # Extract text from image
                ocr_text = self.extract_text_from_image(temp_path)
                
                if ocr_text:
                    images_data.append({
                        'page': i + 1,
                        'ocr_text': ocr_text,
                        'image_path': temp_path
                    })
                
                # Cleanup
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                
        except Exception as e:
            logging.error(f"PDF image extraction failed: {e}")
        
        return images_data
    
    def process_document(self, file_path: str) -> List[DocumentChunk]:
        """Process document with enhanced extraction capabilities."""
        chunks = []
        
        if not os.path.exists(file_path):
            logging.error(f"File not found: {file_path}")
            return chunks
            
        file_ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_ext == '.pdf':
                chunks.extend(self._process_pdf_enhanced(file_path))
            elif file_ext == '.docx':
                chunks.extend(self._process_docx_enhanced(file_path))
            elif file_ext in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
                chunks.extend(self._process_image(file_path))
            elif file_ext in ['.xlsx', '.xls']:
                chunks.extend(self._process_excel(file_path))
            else:
                # Fallback to simple text extraction
                chunks.extend(self._process_text_file(file_path))
                
        except Exception as e:
            logging.error(f"Document processing failed for {file_path}: {e}")
        
        return chunks
    
    def _process_pdf_enhanced(self, pdf_path: str) -> List[DocumentChunk]:
        """Enhanced PDF processing with OCR."""
        chunks = []
        
        try:
            # Regular text extraction
            reader = PdfReader(pdf_path)
            for page_num, page in enumerate(reader.pages):
                text = page.extract_text()
                if text.strip():
                    chunk = DocumentChunk(
                        content=text.strip(),
                        metadata={
                            'page': page_num + 1,
                            'extraction_method': 'text'
                        },
                        chunk_id=f"{os.path.basename(pdf_path)}_page_{page_num + 1}_text",
                        source_file=pdf_path,
                        chunk_index=len(chunks),
                        content_type="text"
                    )
                    chunks.append(chunk)
            
            # OCR for images in PDF (if available)
            if self.ocr_available:
                images_data = self.extract_images_from_pdf(pdf_path)
                for img_data in images_data:
                    if img_data['ocr_text']:
                        chunk = DocumentChunk(
                            content=img_data['ocr_text'],
                            metadata={
                                'page': img_data['page'],
                                'extraction_method': 'ocr'
                            },
                            chunk_id=f"{os.path.basename(pdf_path)}_page_{img_data['page']}_ocr",
                            source_file=pdf_path,
                            chunk_index=len(chunks),
                            content_type="image"
                        )
                        chunks.append(chunk)
                    
        except Exception as e:
            logging.error(f"Enhanced PDF processing failed: {e}")
        
        return chunks
    
    def _process_docx_enhanced(self, docx_path: str) -> List[DocumentChunk]:
        """Enhanced DOCX processing."""
        chunks = []
        
        try:
            doc = Document(docx_path)
            
            # Process text paragraphs
            for i, paragraph in enumerate(doc.paragraphs):
                if paragraph.text.strip():
                    chunk = DocumentChunk(
                        content=paragraph.text.strip(),
                        metadata={'paragraph': i + 1},
                        chunk_id=f"{os.path.basename(docx_path)}_para_{i + 1}",
                        source_file=docx_path,
                        chunk_index=len(chunks),
                        content_type="text"
                    )
                    chunks.append(chunk)
                    
        except Exception as e:
            logging.error(f"Enhanced DOCX processing failed: {e}")
        
        return chunks
    
    def _process_image(self, image_path: str) -> List[DocumentChunk]:
        """Process standalone image with OCR."""
        chunks = []
        
        if not self.ocr_available:
            return chunks
            
        try:
            ocr_text = self.extract_text_from_image(image_path)
            if ocr_text:
                chunk = DocumentChunk(
                    content=ocr_text,
                    metadata={'extraction_method': 'ocr'},
                    chunk_id=f"{os.path.basename(image_path)}_ocr",
                    source_file=image_path,
                    chunk_index=0,
                    content_type="image"
                )
                chunks.append(chunk)
                
        except Exception as e:
            logging.error(f"Image processing failed: {e}")
        
        return chunks
    
    def _process_excel(self, excel_path: str) -> List[DocumentChunk]:
        """Process Excel files."""
        chunks = []
        
        try:
            workbook = openpyxl.load_workbook(excel_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract cell values
                content_rows = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell for cell in row if cell is not None):
                        content_rows.append(' | '.join(str(cell) if cell is not None else '' for cell in row))
                
                if content_rows:
                    content = '\n'.join(content_rows)
                    chunk = DocumentChunk(
                        content=content,
                        metadata={'sheet': sheet_name},
                        chunk_id=f"{os.path.basename(excel_path)}_{sheet_name}",
                        source_file=excel_path,
                        chunk_index=len(chunks),
                        content_type="table"
                    )
                    chunks.append(chunk)
                    
        except Exception as e:
            logging.error(f"Excel processing failed: {e}")
        
        return chunks
    
    def _process_text_file(self, file_path: str) -> List[DocumentChunk]:
        """Process plain text files."""
        chunks = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            if content.strip():
                chunk = DocumentChunk(
                    content=content.strip(),
                    metadata={'file_type': 'text'},
                    chunk_id=f"{os.path.basename(file_path)}_full",
                    source_file=file_path,
                    chunk_index=0,
                    content_type="text"
                )
                chunks.append(chunk)
                
        except Exception as e:
            logging.error(f"Text file processing failed: {e}")
        
        return chunks

class DocumentRAGEngine:
    """Enhanced RAG engine with OCR and advanced document processing."""
    
    def __init__(self, collection_name: str = "qms_documents", persist_directory: str = "./chroma_db"):
        self.collection_name = collection_name
        self.persist_directory = persist_directory
        self.setup_logging()
        
        if not CORE_AVAILABLE:
            self.logger.error("Core RAG dependencies not available!")
            return
            
        self.setup_components()
        self.processor = EnhancedDocumentProcessor()
        
    def setup_logging(self):
        """Setup logging configuration."""
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        
    def setup_components(self):
        """Initialize ChromaDB, embeddings, and AI components."""
        try:
            # Initialize ChromaDB
            os.makedirs(self.persist_directory, exist_ok=True)
            self.client = chromadb.PersistentClient(path=self.persist_directory)
            self.collection = self.client.get_or_create_collection(
                name=self.collection_name,
                metadata={"hnsw:space": "cosine"}
            )
            
            # Initialize embedding model
            self.embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
            
            # Setup Google Gemini
            api_key = os.getenv('GOOGLE_AI_API_KEY', '')
            if api_key:
                genai.configure(api_key=api_key)
                self.model = genai.GenerativeModel('gemini-1.5-flash')
            else:
                self.model = None
                self.logger.warning("No Google AI API key configured")
            
            self.logger.info("Enhanced RAG components initialized successfully")
            
        except Exception as e:
            self.logger.error(f"Setup failed: {e}")
            raise
    
    def index_document(self, file_path: str, metadata: Optional[Dict] = None) -> Dict[str, Any]:
        """Index a document with enhanced processing."""
        try:
            self.logger.info(f"🔄 Processing document: {file_path}")
            
            # Process document with enhanced capabilities
            chunks = self.processor.process_document(file_path)
            
            if not chunks:
                return {"success": False, "message": "No content extracted from document"}
            
            # Generate embeddings and store in ChromaDB
            indexed_chunks = 0
            
            for chunk in chunks:
                try:
                    # Generate embedding
                    embedding = self.embedding_model.encode(chunk.content)
                    # Convert to list for ChromaDB compatibility
                    if hasattr(embedding, 'tolist'):
                        embedding = embedding.tolist()
                    else:
                        embedding = list(embedding.flatten())
                    
                    # Prepare metadata
                    chunk_metadata = {
                        "source_file": chunk.source_file,
                        "chunk_index": chunk.chunk_index,
                        "content_type": chunk.content_type,
                        "indexed_at": datetime.now().isoformat(),
                        **chunk.metadata
                    }
                    
                    if metadata:
                        chunk_metadata.update(metadata)
                    
                    # Add to collection
                    self.collection.add(
                        documents=[chunk.content],
                        embeddings=[embedding],
                        metadatas=[chunk_metadata],
                        ids=[chunk.chunk_id]
                    )
                    
                    indexed_chunks += 1
                    
                except Exception as e:
                    self.logger.warning(f"Failed to index chunk {chunk.chunk_id}: {e}")
            
            result = {
                "success": True,
                "file_path": file_path,
                "chunks_processed": len(chunks),
                "chunks_indexed": indexed_chunks,
                "content_types": list(set(chunk.content_type for chunk in chunks))
            }
            
            self.logger.info(f"✅ Indexed {indexed_chunks}/{len(chunks)} chunks from {file_path}")
            return result
            
        except Exception as e:
            self.logger.error(f"Document indexing failed: {e}")
            return {"success": False, "message": str(e)}
    
    def search_documents(self, query: str, max_results: int = 5) -> List[Dict[str, Any]]:
        """Enhanced semantic search with better ranking."""
        try:
            # Generate query embedding
            query_embedding = self.embedding_model.encode(query)
            if hasattr(query_embedding, 'tolist'):
                query_embedding = query_embedding.tolist()
            elif isinstance(query_embedding, list):
                pass  # already a list
            else:
                query_embedding = query_embedding.flatten().tolist()
            
            # Search in ChromaDB
            results = self.collection.query(
                query_embeddings=[query_embedding],
                n_results=max_results,
                include=['documents', 'metadatas', 'distances']
            )
            
            # Format results with enhanced metadata
            formatted_results = []
            
            if results['documents'] and len(results['documents']) > 0:
                for i in range(len(results['documents'][0])):
                    result = {
                        'content': results['documents'][0][i],
                        'metadata': results['metadatas'][0][i],
                        'similarity': 1 - results['distances'][0][i],  # Convert distance to similarity
                        'content_type': results['metadatas'][0][i].get('content_type', 'text')
                    }
                    formatted_results.append(result)
            
            return formatted_results
            
        except Exception as e:
            self.logger.error(f"Search failed: {e}")
            return []
    
    def delete_document(self, document_id: int, file_path: str = None) -> Dict[str, Any]:
        """
        Löscht alle Vektoren eines Dokuments aus der ChromaDB.
        
        Args:
            document_id (int): ID des Dokuments in der SQL-Datenbank
            file_path (str, optional): Dateipfad des Dokuments
            
        Returns:
            Dict mit Lösch-Statistiken
        """
        try:
            self.logger.info(f"🗑️ Lösche Dokument {document_id} aus RAG-Index...")
            
            # Alle chunk IDs für dieses Dokument finden
            # ChromaDB unterstützt verschiedene Filterstrategien
            deleted_chunks = 0
            
            # Strategie 1: Suche nach document_id in chunk_id Pattern
            if document_id:
                # Get all documents in collection (for filtering)
                all_data = self.collection.get()
                
                # Filter für chunk IDs die zu diesem Dokument gehören
                chunk_ids_to_delete = []
                
                for i, chunk_id in enumerate(all_data['ids']):
                    metadata = all_data['metadatas'][i] if all_data['metadatas'] else {}
                    
                    # Check verschiedene Identifier
                    should_delete = False
                    
                    # Check document_id in metadata
                    if metadata.get('document_id') == document_id:
                        should_delete = True
                    
                    # Check file_path pattern
                    if file_path and metadata.get('source_file'):
                        if file_path in metadata.get('source_file', ''):
                            should_delete = True
                    
                    # Check chunk_id pattern (falls document_id Teil der chunk_id ist)
                    if f"doc_{document_id}_" in chunk_id or f"_{document_id}_" in chunk_id:
                        should_delete = True
                    
                    if should_delete:
                        chunk_ids_to_delete.append(chunk_id)
                
                # Lösche gefundene chunks
                if chunk_ids_to_delete:
                    self.collection.delete(ids=chunk_ids_to_delete)
                    deleted_chunks = len(chunk_ids_to_delete)
                    self.logger.info(f"✅ {deleted_chunks} Chunks gelöscht für Dokument {document_id}")
                else:
                    self.logger.warning(f"⚠️ Keine Chunks gefunden für Dokument {document_id}")
            
            return {
                "success": True,
                "document_id": document_id,
                "deleted_chunks": deleted_chunks,
                "message": f"Dokument {document_id} aus RAG-Index entfernt"
            }
            
        except Exception as e:
            self.logger.error(f"❌ Fehler beim Löschen von Dokument {document_id}: {e}")
            return {
                "success": False,
                "document_id": document_id,
                "deleted_chunks": 0,
                "message": f"Fehler beim Löschen: {str(e)}"
            }
    
    def cleanup_orphaned_vectors(self) -> Dict[str, Any]:
        """
        Bereinigt verwaiste Vektoren (Dokumente die in ChromaDB aber nicht in SQL DB sind).
        
        Returns:
            Dict mit Bereinigungsstatistiken
        """
        try:
            from .database import SessionLocal
            from .models import Document as DocumentModel
            
            self.logger.info("🧹 Starte Bereinigung verwaister Vektoren...")
            
            # Hole alle document_ids aus SQL DB
            db = SessionLocal()
            try:
                sql_document_ids = set()
                sql_file_paths = set()
                
                documents = db.query(DocumentModel).all()
                for doc in documents:
                    sql_document_ids.add(doc.id)
                    if doc.file_path:
                        sql_file_paths.add(doc.file_path)
                        
            finally:
                db.close()
            
            # Hole alle Vektoren aus ChromaDB
            all_data = self.collection.get()
            orphaned_chunks = []
            
            for i, chunk_id in enumerate(all_data['ids']):
                metadata = all_data['metadatas'][i] if all_data['metadatas'] else {}
                
                is_orphaned = False
                
                # Check ob document_id noch existiert
                doc_id = metadata.get('document_id')
                if doc_id and doc_id not in sql_document_ids:
                    is_orphaned = True
                
                # Check ob file_path noch existiert
                source_file = metadata.get('source_file')
                if source_file and source_file not in sql_file_paths:
                    # Zusätzlich prüfen ob Datei physisch existiert
                    if not os.path.exists(source_file):
                        is_orphaned = True
                
                if is_orphaned:
                    orphaned_chunks.append(chunk_id)
            
            # Lösche verwaiste Chunks
            if orphaned_chunks:
                self.collection.delete(ids=orphaned_chunks)
                self.logger.info(f"🧹 {len(orphaned_chunks)} verwaiste Chunks bereinigt")
            else:
                self.logger.info("✅ Keine verwaisten Chunks gefunden")
            
            return {
                "success": True,
                "orphaned_chunks_removed": len(orphaned_chunks),
                "total_chunks_checked": len(all_data['ids']),
                "sql_documents_found": len(sql_document_ids)
            }
            
        except Exception as e:
            self.logger.error(f"❌ Fehler bei Bereinigung: {e}")
            return {
                "success": False,
                "message": str(e)
            }
    
    def chat_with_documents(self, query: str, context_limit: int = 3) -> RAGResult:
        """Enhanced chat with better context handling."""
        try:
            # Search for relevant documents
            search_results = self.search_documents(query, max_results=context_limit)
            
            if not search_results:
                return RAGResult(
                    response="Keine relevanten Dokumente gefunden.",
                    sources=[],
                    confidence=0.0
                )
            
            # Build enhanced context
            context_parts = []
            sources = []
            
            for result in search_results:
                content_type = result['content_type']
                source_info = result['metadata'].get('source_file', 'Unbekannt')
                
                context_part = f"[{content_type.upper()}] {result['content']}"
                context_parts.append(context_part)
                
                sources.append({
                    'file': os.path.basename(source_info),
                    'content_type': content_type,
                    'similarity': result['similarity'],
                    'metadata': result['metadata']
                })
            
            context = "\n\n".join(context_parts)
            
            # Enhanced prompt
            prompt = f"""Du bist ein QMS-Experte. Beantworte die Frage basierend auf dem bereitgestellten Kontext.

KONTEXT:
{context}

FRAGE: {query}

ANWEISUNG:
- Antworte präzise und fachlich korrekt
- Beziehe dich explizit auf die Quelldokumente
- Erwähne den Inhaltstyp (Text, Bild, Tabelle) wenn relevant
- Wenn unsicher, sage es explizit

ANTWORT:"""

            # Generate response
            if self.model:
                response = self.model.generate_content(prompt)
                response_text = response.text
            else:
                response_text = f"KI-Antwort nicht verfügbar. Gefundene Quellen:\n\n{context}"
            
            # Calculate confidence based on similarity scores
            avg_confidence = sum(r['similarity'] for r in search_results) / len(search_results)
            
            return RAGResult(
                response=response_text,
                sources=sources,
                confidence=avg_confidence
            )
            
        except Exception as e:
            self.logger.error(f"Chat failed: {e}")
            return RAGResult(
                response=f"Fehler bei der Antwortgenerierung: {str(e)}",
                sources=[],
                confidence=0.0
            )
    
    def get_system_status(self) -> Dict[str, Any]:
        """Get enhanced system status."""
        try:
            collection_count = self.collection.count()
            
            # Get content type distribution
            all_results = self.collection.get(include=['metadatas'])
            content_types = {}
            
            if all_results['metadatas']:
                for metadata in all_results['metadatas']:
                    content_type = metadata.get('content_type', 'unknown')
                    content_types[content_type] = content_types.get(content_type, 0) + 1
            
            return {
                "status": "operational",
                "documents_indexed": collection_count,
                "content_types": content_types,
                "embedding_model": "all-MiniLM-L6-v2",
                "ai_model": "gemini-1.5-flash",
                "ocr_enabled": OCR_AVAILABLE,
                "core_available": CORE_AVAILABLE,
                "supported_formats": ["PDF", "DOCX", "Images", "Excel", "Text"]
            }
            
        except Exception as e:
            return {"status": "error", "message": str(e)}

# Global RAG engine instance
rag_engine = None

def get_rag_engine() -> Optional[DocumentRAGEngine]:
    """Get global RAG engine instance."""
    global rag_engine
    if rag_engine is None and CORE_AVAILABLE:
        try:
            rag_engine = DocumentRAGEngine()
        except Exception as e:
            logging.error(f"Failed to initialize RAG engine: {e}")
            return None
    return rag_engine

async def index_all_documents(db):
    """Index all documents in database for RAG."""
    engine = get_rag_engine()
    if not engine:
        return {"success": False, "message": "RAG engine not available"}
    
    # This would need to be adapted based on your actual document model
    # For now, return a placeholder
    return {"success": True, "message": "Document indexing would be implemented here"} 